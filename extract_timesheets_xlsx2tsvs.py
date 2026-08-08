#!/usr/bin/env python3
"""Extract per-employee timesheets for a given month from the attendance xlsx.

Reads an AttendanceYYYY.xlsx (one tab per employee, named like Beth_1) and
writes one TSV per tab into a destination directory, filtered to the
requested month. The tab name becomes the TSV filename. The TSV columns
match what src.loaders.load_timesheet_folder expects.

run_payroll.py calls extract_month() directly against its temp workdir; the
CLI below is for inspecting a downloaded workbook by hand.

Usage:
    python extract_timesheets_xlsx2tsvs.py --xlsx Attendance2026.xlsx \\
        --dest ./ts_2026_04 --year 2026 --month 4
"""

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


# The first two columns are read positionally: their header cells are
# unreliable across tabs (variously "date", "Date", blank, or a stray number).
# Everything after is matched by header name, because the optional columns do
# not sit at the same index on every tab -- notes is column 8 on one tab and 9
# on another, so reading by position would silently mix them up.
HEADER = [
    "date", "wkdy", "hrs_norm", "hrs_wrkd",
    "hrs_miss", "hrs_sik", "hrs_ot_1_5", "hrs_ot_2_0",
    # One-off money adjustments for the month, e.g. a completion bonus or a
    # correction for holidays worked in an earlier period. Both are taxable;
    # they differ only in whether the 15% housing allowance applies on top.
    "adj_with_housing", "adj_no_housing",
    "notes",
]

_BY_NAME = HEADER[2:]


def _fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v if v is not None else ""


def _fmt_wkdy(v):
    if isinstance(v, datetime):
        return v.strftime("%a")
    return v if v is not None else ""


def _fmt_text(v):
    return "" if v is None else str(v).strip()


def _fmt_num(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def extract_month(xlsx_path: Path, dest_dir: Path, year: int, month: int,
                  log=print) -> int:
    """Extract one TSV per worksheet, filtered to the given year/month.

    Returns the number of TSV files written.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dest_dir.mkdir(parents=True, exist_ok=True)
    # Clear any stale TSVs from prior runs (e.g. renamed sheets)
    for stale in dest_dir.glob("*.tsv"):
        stale.unlink()

    written = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [str(c).strip() if c is not None else ""
                  for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        index = {name: i for i, name in enumerate(header) if name}

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = row[0]
            if not isinstance(d, datetime):
                continue
            if d.year != year or d.month != month:
                continue

            def cell(name):
                i = index.get(name)
                return row[i] if i is not None and i < len(row) else None

            rows.append(
                [_fmt_date(d), _fmt_wkdy(row[1])]
                + [_fmt_text(cell(n)) if n == "notes" else _fmt_num(cell(n))
                   for n in _BY_NAME]
            )

        out_path = dest_dir / f"{sheet_name}.tsv"
        with open(out_path, "w", newline="") as f:
            writer = csv.writer(f, delimiter="\t")
            writer.writerow(HEADER)
            writer.writerows(rows)
        written += 1
        log(f"  {sheet_name}.tsv: {len(rows)} rows")

    return written


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument("--xlsx", type=Path, required=True,
                        help="Path to AttendanceYYYY.xlsx")
    parser.add_argument("--dest", type=Path, required=True,
                        help="Directory to write per-employee TSVs into")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--month", type=int, default=3)
    args = parser.parse_args()

    xlsx_path = args.xlsx
    if not xlsx_path.is_file():
        print(f"Source not found: {xlsx_path}", file=sys.stderr)
        return 1

    dest_dir = args.dest
    print(f"Extracting {args.year}-{args.month:02d} from {xlsx_path}")
    print(f"Writing to {dest_dir}")
    print()

    written = extract_month(xlsx_path, dest_dir, args.year, args.month)
    print()
    print(f"Wrote {written} TSVs to {dest_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
